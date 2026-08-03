"""
Parser de los Sheets de "vaciado" viejos (formato: una hoja por mes + RESUMEN +
CONTRACT LABOR) hacia el esquema de ContabilidadMensualHistorica.

Diseñado para ser TOLERANTE a variaciones entre archivos de ~110 compañías
distintas hechos por distintas personas a lo largo de años — busca los
encabezados por contenido, no por posición fija de fila/columna, y devuelve
advertencias en vez de fallar cuando algo no calza con el patrón esperado.

Solo lee la hoja RESUMEN (no las hojas de mes con el detalle transacción por
transacción) — el usuario confirmó que para el histórico solo hacen falta los
totales, no cada gasto.
"""
from __future__ import annotations

from typing import Any, Optional

import openpyxl

MESES = {
    "ENERO": 1, "FEBRERO": 2, "MARZO": 3, "ABRIL": 4, "MAYO": 5, "JUNIO": 6,
    "JULIO": 7, "AGOSTO": 8, "SEPTIEMBRE": 9, "OCTUBRE": 10, "NOVIEMBRE": 11, "DICIEMBRE": 12,
}

# Nombres de categoría tal cual aparecen en los Sheets reales (ver categoria_gasto
# seed en contabilidad.py) — usado para reconocer filas de categoría en el bloque
# de "GASTOS por mes" de RESUMEN, sin depender de su posición exacta.
NOMBRES_CATEGORIA_CONOCIDOS = {
    "ADVERTISING", "CAR EXPENSES", "COMISION/FEES", "COMMISSION/FEES", "CONTRACT LABOR",
    "DEPRECIATION", "INSURANCE", "LEGAL SERVICES", "OFFICE EXPENSES", "RENT/LEASE",
    "REPAIRS/MAINT", "SUPPLIES", "TAXES/LICENSES", "TRAVEL", "MEALS", "UTILITIES",
    "OTHER", "OTHER/CHEQUES", "OTHER/ CHEQUES", "PAGOS A TARJETAS", "PAGOS A TAREJTAS",
    "ATM", "PERSONAL", "PERSONAL/ATM", "PERSONAL/ atm", "SAVINGS", "CHEQUES",
}

# Normaliza variantes de escritura encontradas en archivos reales a un nombre único.
ALIAS_CATEGORIA = {
    "COMMISSION/FEES": "COMISION/FEES",
    "OTHER/CHEQUES": "OTHER",
    "OTHER/ CHEQUES": "OTHER",
    "PAGOS A TAREJTAS": "PAGOS A TARJETAS",  # typo real encontrado en un archivo (Solorio)
    "PERSONAL/ATM": "PERSONAL",
    "PERSONAL/ atm": "PERSONAL",
}


def _norm(v: Any) -> str:
    return str(v).strip().upper() if v is not None else ""


def _num(v: Any) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace("$", "").replace(",", "").strip() or 0)
    except ValueError:
        return 0.0


def _normaliza_categoria(nombre: str) -> str:
    n = _norm(nombre)
    return ALIAS_CATEGORIA.get(n, n)


class VaciadoParseResult:
    def __init__(self):
        self.meses: dict[int, dict] = {}  # mes(1-12) -> {ingreso, gasto_deducible, saldo_final, categorias: {}}
        self.advertencias: list[str] = []

    def mes(self, m: int) -> dict:
        return self.meses.setdefault(m, {
            "ingreso_total": 0.0, "saldo_final": None, "gasto_por_categoria": {},
        })


def _parse_resumen_totales_mensuales(ws, resultado: VaciadoParseResult) -> None:
    """Bloque 'TOTAL | DEPOSITOS | GASTOS | | SALDO FINAL' por mes."""
    header_row = None
    for r in range(1, min(ws.max_row, 200) + 1):
        vals = [_norm(ws.cell(row=r, column=c).value) for c in range(1, 6)]
        if vals[0] == "TOTAL" and vals[1] == "DEPOSITOS" and "GASTOS" in vals[2]:
            header_row = r
            break
    if header_row is None:
        resultado.advertencias.append("No se encontró el bloque 'TOTAL/DEPOSITOS/GASTOS' en RESUMEN.")
        return

    r = header_row + 1
    seen = 0
    while r <= ws.max_row and seen < 12:
        nombre_mes = _norm(ws.cell(row=r, column=1).value)
        if nombre_mes in MESES:
            m = MESES[nombre_mes]
            deposito = _num(ws.cell(row=r, column=2).value)
            saldo_final = ws.cell(row=r, column=5).value
            entry = resultado.mes(m)
            entry["ingreso_total"] = deposito
            entry["saldo_final"] = _num(saldo_final) if saldo_final is not None else None
            seen += 1
        elif nombre_mes == "TOTAL":
            break
        r += 1


def _parse_resumen_categorias(ws, resultado: VaciadoParseResult) -> None:
    """Bloques 'GASTOS | ENERO | FEBRERO | ...' (normalmente 2 bloques de 6
    meses) — busca TODAS las ocurrencias, no asume una posición fija."""
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            if _norm(ws.cell(row=r, column=c).value) != "GASTOS":
                continue
            # ¿Los encabezados a la derecha son nombres de mes? si no, no es este bloque.
            month_cols: dict[int, int] = {}  # columna -> mes(1-12)
            cc = c + 1
            while cc <= ws.max_column:
                v = _norm(ws.cell(row=r, column=cc).value)
                if v in MESES:
                    month_cols[cc] = MESES[v]
                    cc += 1
                else:
                    break
            if not month_cols:
                continue  # es el "GASTOS" del bloque de totales, no de categorías

            # Filas de categoría debajo, hasta la primera fila que no sea una
            # categoría conocida (ej. fila en blanco, o el siguiente bloque).
            rr = r + 1
            while rr <= ws.max_row:
                nombre_cat = ws.cell(row=rr, column=c).value
                if nombre_cat is None or _norm(nombre_cat) not in NOMBRES_CATEGORIA_CONOCIDOS:
                    break
                cat = _normaliza_categoria(nombre_cat)
                for col, mes in month_cols.items():
                    monto = _num(ws.cell(row=rr, column=col).value)
                    if monto:
                        entry = resultado.mes(mes)
                        entry["gasto_por_categoria"][cat] = entry["gasto_por_categoria"].get(cat, 0.0) + monto
                rr += 1


def parse_vaciado(path: str, categorias_deducibles: set[str]) -> VaciadoParseResult:
    """Punto de entrada: abre el workbook, lee RESUMEN, devuelve un resultado
    con los totales por mes. `categorias_deducibles` = set de nombres de
    categoría que SÍ cuentan en el gasto total (todas menos las de control)."""
    resultado = VaciadoParseResult()
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:  # noqa: BLE001
        resultado.advertencias.append(f"No se pudo abrir el archivo: {e}")
        return resultado

    if "RESUMEN" not in wb.sheetnames:
        resultado.advertencias.append("El archivo no tiene una hoja 'RESUMEN'.")
        return resultado

    ws = wb["RESUMEN"]
    _parse_resumen_totales_mensuales(ws, resultado)
    _parse_resumen_categorias(ws, resultado)

    for mes, entry in resultado.meses.items():
        gasto_deducible = sum(
            monto for cat, monto in entry["gasto_por_categoria"].items()
            if cat in categorias_deducibles
        )
        entry["gasto_total_deducible"] = round(gasto_deducible, 2)

    # Descarta meses completamente vacíos (compañías con menos de 12 meses de datos)
    resultado.meses = {
        m: e for m, e in resultado.meses.items()
        if e["ingreso_total"] or e["gasto_por_categoria"] or e.get("saldo_final")
    }
    if not resultado.meses:
        resultado.advertencias.append("No se encontraron meses con datos (¿archivo en blanco?).")

    return resultado
