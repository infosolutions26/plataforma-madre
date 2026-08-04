"""
Parser de las hojas MENSUALES (ENERO..DICIEMBRE) de los Sheets de vaciado en
Drive — las que tienen el detalle transacción por transacción, NO la hoja
RESUMEN (esa la lee vaciado_parser.py y solo trae totales).

Layout real encontrado (Solorio, Enero 2026): las categorías van como
encabezados horizontales en una fila (columnas pares), y debajo de cada una
sus transacciones en un par de columnas: el MONTO en la columna de la
categoría y el COMERCIO en la columna de la derecha. Ej.:

    col 4 = CAR EXPENSES        col 5 = (descripción)
       36.43                       INK #3218 SNK FUEL
       52.57                       INK #3218 SNK FUEL

De aquí sale la data de entrenamiento del diccionario Comercio: pares
(texto_crudo_del_comercio -> categoría) para que el sistema nuevo sugiera
categoría automáticamente. `normaliza_comercio` limpia el ruido de
tarjeta/cuenta ('INK #3218', 'COMPL', '#NNNN', 'SQ *', etc.) para que el
mismo comercio de distintos clientes/meses colapse en una sola entrada.
"""
from __future__ import annotations

import re
import unicodedata
from typing import Any

import openpyxl

MESES = {
    "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO",
    "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE",
}

# Encabezados de categoría tal cual aparecen en las hojas de mes (difieren un
# poco de los de la hoja RESUMEN) -> nombre canónico del seed CategoriaGasto.
HEADER_A_CATEGORIA = {
    "ADVERTISING": "ADVERTISING",
    "CAR EXPENSES": "CAR EXPENSES",
    "COMISION/FEES": "COMISION/FEES",
    "COMMISSION/FEES": "COMISION/FEES",
    "CONTRACT LABOR": "CONTRACT LABOR",
    "DEPRECIATION": "DEPRECIATION",
    "INSURANCE": "INSURANCE",
    "LEGAL SERVICES": "LEGAL SERVICES",
    "OFFICE EXPENSES": "OFFICE EXPENSES",
    "RENT/LEASE": "RENT/LEASE",
    "REPAIRS/MAINT": "REPAIRS/MAINT",
    "SUPPLIES": "SUPPLIES",
    "TAXES/LICENSES": "TAXES/LICENSES",
    "TRAVEL": "TRAVEL",
    "MEALS": "MEALS",
    "UTILITIES": "UTILITIES",
    "OTHER": "OTHER",
    "OTHER/CHEQUES": "OTHER",
    # categorías de control (se guardan pero no entran a las tablas top-10 del
    # entregable; Contract Labor sí es útil, el resto casi no):
    "TARJETAS": "PAGOS A TARJETAS",
    "PAGOS A TARJETAS": "PAGOS A TARJETAS",
    "ATM": "ATM",
    "ATM/PERSONAL": "PERSONAL",
    "PERSONAL": "PERSONAL",
    "PERSONAL/ATM": "PERSONAL",
    "SAVINGS": "SAVINGS",
    "CHEQUES": "CHEQUES",
}

# Filas de plantilla que hay que ignorar (la hoja trae ejemplos precargados).
COMERCIO_BASURA = {"EJEMPLO", "EJEMPLOS", "TOTALES", "TOTAL", "SALDO INICIAL",
                   "SALDO FINAL", "DEPOSITOS", "DEPÓSITOS", "GASTOS TOTALES", ""}

# Ruido de tarjeta/cuenta al inicio de la descripción, que NO es el comercio.
_PREFIJOS_RUIDO = re.compile(
    r"^(INK(\s+COMPL)?|COMPL|DEBIT|CREDIT|POS|PURCHASE|PYMT|PAYMENT|RECURRING|"
    r"CHASE|BOFA|WELLS\s*FARGO|WF|CITI|AMEX|CAPITAL\s*ONE|CAP1|VISA|MASTERCARD|MC|"
    r"SQ|TST|PP|PAYPAL|GOOGLE|GOOGLE\s*\*|GPAY|APLPAY|APPLE\s*PAY)\b[\s*#:.-]*",
    re.IGNORECASE,
)
_TAG_CUENTA = re.compile(r"#\s*\d{3,}")           # '#3218', '# 1801'
_ASTERISCO = re.compile(r"[*]+")


def _num(v: Any) -> float | None:
    if isinstance(v, (int, float)):
        return float(v)
    if v is None:
        return None
    try:
        return float(str(v).replace("$", "").replace(",", "").strip())
    except ValueError:
        return None


def normaliza_comercio(texto: str) -> str:
    """Limpia la descripción cruda a una clave de comercio estable. Conservador
    a propósito: si sobra ruido, el encargado lo edita después (nombre_editado);
    lo importante es que 'INK #3218 HOME DEPOT' y 'HOME DEPOT' colapsen igual."""
    s = str(texto).upper().strip()
    # quita acentos para que 'JOSÉ' y 'JOSE' colapsen (son el mismo colaborador)
    s = "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
    s = _TAG_CUENTA.sub(" ", s)
    # aplica el barrido de prefijos de ruido repetidas veces (ej. 'INK COMPL #x HOME DEPOT')
    for _ in range(3):
        nuevo = _PREFIJOS_RUIDO.sub("", s).strip()
        if nuevo == s:
            break
        s = nuevo
    s = _ASTERISCO.sub(" ", s)
    s = re.sub(r"^THE\s+", "", s)  # 'THE HOME DEPOT' -> 'HOME DEPOT'
    s = re.sub(r"\s+", " ", s).strip(" -.*#")
    return s


def _fila_encabezados(ws) -> int | None:
    """Encuentra la fila que trae los encabezados de categoría (la que más
    nombres conocidos contiene), sin asumir que es la fila 2."""
    mejor_fila, mejor_conteo = None, 0
    for r in range(1, min(ws.max_row, 15) + 1):
        conteo = 0
        for c in range(1, min(ws.max_column, 60) + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None and str(v).strip().upper() in HEADER_A_CATEGORIA:
                conteo += 1
        if conteo > mejor_conteo:
            mejor_fila, mejor_conteo = r, conteo
    return mejor_fila if mejor_conteo >= 3 else None


def parse_hoja_mes(ws) -> list[tuple[float, str, str]]:
    """Devuelve [(monto, comercio_raw, categoria_canonica), ...] de una hoja de mes."""
    fila_hdr = _fila_encabezados(ws)
    if fila_hdr is None:
        return []
    # columna de categoría -> categoría canónica
    cols_cat: dict[int, str] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=fila_hdr, column=c).value
        if v is not None:
            cat = HEADER_A_CATEGORIA.get(str(v).strip().upper())
            if cat:
                cols_cat[c] = cat

    resultados: list[tuple[float, str, str]] = []
    for col, cat in cols_cat.items():
        for r in range(fila_hdr + 1, ws.max_row + 1):
            monto = _num(ws.cell(row=r, column=col).value)
            comercio = ws.cell(row=r, column=col + 1).value
            if monto is None or not monto or comercio is None:
                continue
            texto = str(comercio).strip()
            if texto.upper() in COMERCIO_BASURA:
                continue
            resultados.append((monto, texto, cat))
    return resultados


def parse_workbook_mensual(path: str) -> list[tuple[float, str, str]]:
    """Recorre todas las hojas de mes de un workbook y junta sus transacciones."""
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception:  # noqa: BLE001
        return []
    todo: list[tuple[float, str, str]] = []
    for nombre in wb.sheetnames:
        if nombre.strip().upper() in MESES:
            todo.extend(parse_hoja_mes(wb[nombre]))
    return todo
